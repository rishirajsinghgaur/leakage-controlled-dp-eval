r"""Membership-inference (LiRA) on the SWaT December-2015 LONG normal stream.

Purpose: close the July-2019 limitation that the blocked split yields only ~8 blocks.
The Dec-2015 normal record is a contiguous 7-day 1 Hz stream (~495k rows); using a long
prefix gives many more blocks, so the blocked-split MIA is no longer small-sample fragile.

Identical attack to experiments/mia_replicate.py (offline one-sided LiRA, R_REF reference
models, positive control = non-private overfit). Only the stream length / block geometry
change. Resume-safe: skips (target,eps,split,seed) cells already in the JSON, so a monitor
loop can restart it until complete. Writes results/mia_swat2015.json(+_summary.json).
"""
import sys, json, logging
from pathlib import Path
import numpy as np
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT)); sys.stdout.reconfigure(encoding="utf-8")
logging.basicConfig(level=logging.ERROR)
import torch, torch.nn as nn
from torch.utils.data import DataLoader, TensorDataset
from sklearn.metrics import roc_auc_score
from sklearn.preprocessing import StandardScaler
from models.mlp import AnomalyAutoencoder
from privacy.accountant import compute_sigma_for_total_epsilon
from experiments.run_full_paper_sweep import DS_CFG

NORMAL_XLSX = ROOT.parent / "Dataset" / "swat" / "SWaT.A1 & A2_Dec 2015" / "Physical" / "SWaT_Dataset_Normal_v1.xlsx"
CACHE = ROOT / "results" / "cache_swat2015_normal.npz"
OUT = ROOT / "results" / "mia_swat2015.json"
SUMM = ROOT / "results" / "mia_swat2015_summary.json"

NMAX_NORMAL = 100_000         # contiguous prefix of the 7-day normal stream (~28 h at 1 Hz)
BLOCK = 2500; GAP = 1000      # gap > Dec-2015 autocorr length (median ~920); ~40 blocks
R_REF = 6; OVERFIT_PASSES = 80; DP_PASSES = 12
EPS_LIST = [0.5, 1.0, 2.0, 4.0]
SEEDS = [200, 201, 202, 203]
SPLITS = ["contiguous", "random", "blocked"]
CONFIGS = [("np_overfit", None)] + [("dp", e) for e in EPS_LIST]


def load_normal_stream():
    if CACHE.exists():
        return np.load(CACHE)["X"].astype("float32")
    import openpyxl
    print(f"Streaming first {NMAX_NORMAL} rows of {NORMAL_XLSX.name}...", flush=True)
    wb = openpyxl.load_workbook(NORMAL_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, max_row=2 + NMAX_NORMAL, values_only=True)  # row2 header
    header = [str(c).strip() for c in next(rows)]
    ts_i = next(i for i, h in enumerate(header) if "time" in h.lower())
    lab_i = next(i for i, h in enumerate(header) if h.lower().replace(" ", "") == "normal/attack")
    data = []
    for r in rows:
        data.append([r[i] for i in range(len(header)) if i not in (ts_i, lab_i)])
    wb.close()
    arr = np.array(data, dtype=np.float32)
    keep = arr.std(0) > 0                              # drop constant tags (same filter as loaders)
    arr = arr[:, keep]
    X = StandardScaler().fit_transform(arr).astype(np.float32)
    CACHE.parent.mkdir(exist_ok=True)
    np.savez_compressed(CACHE, X=X)
    print(f"Cached normal stream: {X.shape[0]} rows x {X.shape[1]} tags", flush=True)
    return X


def autocorr_len(X, thresh=1.0 / np.e, maxlag=2000):
    """Mean over tags of the lag at which |autocorrelation| first drops below 1/e."""
    lags = []
    for j in range(X.shape[1]):
        x = X[:20000, j] - X[:20000, j].mean()
        denom = (x * x).sum()
        if denom <= 0:
            continue
        for L in range(1, maxlag):
            ac = (x[:-L] * x[L:]).sum() / denom
            if abs(ac) < thresh:
                lags.append(L); break
        else:
            lags.append(maxlag)
    return float(np.median(lags)) if lags else 0.0


def train_ae(X, cfg, dp_sigma=None, seed=0, passes=12):
    torch.manual_seed(seed); np.random.seed(seed)
    n = len(X); B = min(cfg["batch_size"], max(2, n - 1))
    m = AnomalyAutoencoder(X.shape[1], cfg["bottleneck"]); opt = torch.optim.Adam(m.parameters(), lr=cfg["lr"])
    ld = DataLoader(TensorDataset(torch.from_numpy(X).float()), batch_size=B, shuffle=True, drop_last=True)
    if dp_sigma is not None:
        from opacus import PrivacyEngine
        m, opt, ld = PrivacyEngine().make_private(module=m, optimizer=opt, data_loader=ld,
                                                  noise_multiplier=dp_sigma, max_grad_norm=cfg["max_grad_norm"])
    lf = nn.MSELoss(); m.train()
    for _ in range(passes):
        for (b,) in ld:
            opt.zero_grad(); lf(m(b), b).backward(); opt.step()
    return getattr(m, "_module", m)


def losses(model, X):
    model.eval()
    with torch.no_grad():
        r = model(torch.from_numpy(X).float()); return ((r - torch.from_numpy(X).float()) ** 2).mean(1).numpy()


def split_idx(n, split, rng):
    if split == "contiguous":
        nh = n // 2; return np.arange(nh), np.arange(nh, 2 * nh), 1
    if split == "random":
        p = rng.permutation(n); nh = n // 2; return p[:nh], p[nh:2 * nh], 0
    nb = n // BLOCK; a = rng.integers(0, 2, nb); mem = []; non = []; used = 0
    for b in range(nb):
        s = b * BLOCK + GAP; e = (b + 1) * BLOCK - GAP
        if e <= s: continue
        used += 1; (mem if a[b] else non).extend(range(s, e))
    return np.array(mem, int), np.array(non, int), used


def attack(Xn, cfg, kind, eps, split, seed):
    rng = np.random.default_rng(seed); n = len(Xn)
    mi, ni, nblocks = split_idx(n, split, rng); m = min(len(mi), len(ni))
    if m < 40: return None
    mi = mi[:m]; ni = ni[:m]; idx = np.concatenate([mi, ni]); Xs = Xn[idx]
    label = np.zeros(len(Xs), int); label[:m] = 1; Xmem = Xn[mi]
    if kind == "dp":
        B = min(cfg["batch_size"], len(Xmem) - 1); steps = max(1, len(Xmem) // B) * DP_PASSES
        sig = compute_sigma_for_total_epsilon(eps, len(Xmem), B, steps, 1, cfg["delta"])
        tgt = train_ae(Xmem, cfg, dp_sigma=sig, seed=seed, passes=DP_PASSES)
    else:
        tgt = train_ae(Xmem, cfg, dp_sigma=None, seed=seed, passes=OVERFIT_PASSES)
    Ltgt = losses(tgt, Xs)
    ref_in = [[] for _ in range(len(Xs))]
    for r in range(R_REF):
        ridx = rng.permutation(len(Xs)); half = ridx[:len(Xs) // 2]
        refm = train_ae(Xs[half], cfg, dp_sigma=None, seed=5000 + 10 * seed + r, passes=DP_PASSES)
        L = losses(refm, Xs); inset = set(half.tolist())
        for i in range(len(Xs)):
            if i not in inset: ref_in[i].append(L[i])
    score = np.zeros(len(Xs))
    for i in range(len(Xs)):
        o = np.array(ref_in[i]) if ref_in[i] else np.array([Ltgt.mean()])
        mu, sd = o.mean(), o.std() + 1e-6; score[i] = (mu - Ltgt[i]) / sd
    return float(roc_auc_score(label, score)), int(m), int(nblocks)


def main():
    calib = "calib" in sys.argv
    cfg = DS_CFG["swat"]
    Xn = load_normal_stream()
    acl = autocorr_len(Xn)
    print(f"[swat2015] normal n={len(Xn)}, tags={Xn.shape[1]}, autocorr_len(median)~{acl:.0f}, "
          f"GAP={GAP} ({'OK gap>acl' if GAP >= acl else 'WARN gap<acl'}), block={BLOCK}", flush=True)
    out = json.loads(OUT.read_text()) if OUT.exists() else []
    done = {(r["target"], r["eps"], r["split"], r["seed"]) for r in out}
    summ = json.loads(SUMM.read_text()) if SUMM.exists() else []
    configs = [("np_overfit", None), ("dp", 2.0)] if calib else CONFIGS
    splits = ["blocked"] if calib else SPLITS
    seeds = SEEDS[:1] if calib else SEEDS
    for kind, eps in configs:
        for split in splits:
            aucs = []; nm = None; nb = None
            for seed in seeds:
                if (kind, eps, split, seed) in done:
                    prev = next(r for r in out if (r["target"], r["eps"], r["split"], r["seed"]) == (kind, eps, split, seed))
                    aucs.append(prev["lira"]); nm = prev["n_member"]; nb = prev.get("n_blocks")
                    continue
                res = attack(Xn, cfg, kind, eps, split, seed)
                if res is None: break
                a, nm, nb = res; aucs.append(a)
                out.append({"dataset": "swat2015", "target": kind, "eps": eps, "split": split,
                            "seed": seed, "lira": round(a, 4), "n_member": nm, "n_blocks": nb})
                OUT.write_text(json.dumps(out, indent=2))
                print(f"  {kind}{'@'+str(eps) if eps else ''}/{split} seed={seed}: "
                      f"LiRA={a:.4f} (n_member={nm}, blocks={nb})", flush=True)
            if aucs:
                s = {"dataset": "swat2015", "target": kind, "eps": eps, "split": split,
                     "lira_mean": round(float(np.mean(aucs)), 4), "lira_std": round(float(np.std(aucs)), 4),
                     "seeds": len(aucs), "n_member": nm, "n_blocks": nb}
                summ = [x for x in summ if not (x["target"] == kind and x["eps"] == eps and x["split"] == split)]
                summ.append(s); SUMM.write_text(json.dumps(summ, indent=2))
                print(f"  -> {kind}{'@'+str(eps) if eps else ''}/{split}: "
                      f"LiRA {s['lira_mean']}+/-{s['lira_std']} (blocks={nb})", flush=True)
    print("\nDONE mia_swat2015.", flush=True)


if __name__ == "__main__":
    main()
